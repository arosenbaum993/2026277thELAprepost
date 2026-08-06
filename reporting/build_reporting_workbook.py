#!/usr/bin/env python3
"""
Build the Grade 7 ELA Pre/Post Results Reporting Workbook.

Generates GA7-ELA-PrePost-Reporting-Workbook.xlsx with:
  START HERE · Settings · Roster · Pre-Test Entry · Post-Test Entry ·
  Student Report · Class Dashboard · Item Analysis · Intervention Groups ·
  Standards Coverage

Every computed cell is a live Excel formula, so the workbook recalculates
when a teacher edits scores, cut points, or thresholds.

Usage:  python build_reporting_workbook.py [output.xlsx] [roster_capacity]

roster_capacity defaults to 150 (five sections of 30). Lowering it produces a
smaller, faster-opening file; raising it adds rows. Every formula scales with it.
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

OUT = sys.argv[1] if len(sys.argv) > 1 else "GA7-ELA-PrePost-Reporting-Workbook.xlsx"

FIRST_ROW = 6            # first student data row (same on every sheet)
N_STUDENTS = int(sys.argv[2]) if len(sys.argv) > 2 else 150   # roster capacity
LAST_ROW = FIRST_ROW + N_STUDENTS - 1     # 155

# Item specification: (label, max_points, category, standard, expectation, DOK, trait)
ITEMS = [
    ("Q1",  1, "I-TEC", "7.T.T.1",  "7.T.T.1.a",           2, ""),
    ("Q2",  2, "I-TEC", "7.T.T.1",  "7.T.T.1.b",           3, ""),
    ("Q3",  1, "I-LNG", "7.L.V.3",  "7.L.V.3.b",           2, ""),
    ("Q4",  1, "I-CSS", "7.T.SS.2", "7.T.SS.2.a",          2, ""),
    ("Q5",  1, "I-TEC", "7.T.T.1",  "7.T.T.1.c",           2, ""),
    ("Q6",  1, "I-TEC", "7.T.T.4",  "7.T.T.4.a",           1, ""),
    ("Q7",  2, "I-TEC", "7.T.T.4",  "7.T.T.4.a",           3, ""),
    ("Q8",  2, "I-TEC", "7.T.T.1",  "7.T.T.1.c",           3, ""),
    ("Q9",  1, "I-PRA", "7.T.PM.1", "7.T.PM.1.a",          2, ""),
    ("Q10", 2, "I-PRA", "7.T.T.1",  "7.T.T.1.d/7.T.PM.1.a", 3, ""),
    ("Q11", 1, "I-CSS", "7.T.SS.1", "7.T.SS.1.a",          2, ""),
    ("Q12", 1, "I-TEC", "7.T.T.2",  "7.T.T.2.a",           2, ""),
    ("Q13", 1, "I-TEC", "7.T.T.2",  "7.T.T.2.a",           2, ""),
    ("Q14", 1, "I-LNG", "7.L.V.2",  "7.L.V.2.a",           1, ""),
    ("Q15", 1, "I-LNG", "7.L.V.2",  "7.L.V.2.b",           2, ""),
    ("Q16", 2, "I-CSS", "7.T.C.1",  "7.T.C.1.a",           3, ""),
    ("Q17", 1, "I-TEC", "7.T.T.3",  "7.T.T.3.a",           2, ""),
    ("Q18", 2, "I-TEC", "7.T.T.3",  "7.T.T.3.a",           3, ""),
    ("Q19", 1, "I-LNG", "7.L.V.3",  "7.L.V.3.c",           2, ""),
    ("Q20", 2, "I-CSS", "7.T.C.2",  "7.T.C.2.a/b",         3, ""),
    ("Q21", 2, "I-LNG", "7.L.GC.2", "7.L.GC.2.a",          2, ""),
    ("Q22", 2, "I-TEC", "7.T.T.2",  "7.T.T.2.b",           3, ""),
    ("Q23", 1, "I-PRA", "7.T.RA.2", "7.T.RA.2.b",          2, ""),
    ("Q24", 2, "I-PRA", "7.T.RA.2", "7.T.RA.2.a",          2, ""),
    ("Q25", 1, "C-CSS", "7.T.SS.1", "7.T.SS.1.d",          2, "1"),
    ("Q26", 1, "C-TEC", "7.T.T.3",  "7.T.T.3.c",           2, "2"),
    ("Q27", 2, "C-TEC", "7.T.T.3",  "7.T.T.3.c/7.T.RA.2.c", 3, "2"),
    ("Q28", 2, "C-TEC", "7.T.T.2",  "7.T.T.2.d",           2, "2"),
    ("Q29", 1, "C-TEC", "7.T.T.3",  "7.T.T.3.c",           2, "2"),
    ("Q30", 2, "C-CSS", "7.T.SS.1", "7.T.SS.1.c",          3, "1"),
    ("Q31", 1, "C-CSS", "7.T.SS.2", "7.T.SS.2.c",          2, "1"),
    ("Q32", 1, "C-LNG", "7.L.GC.2", "7.L.GC.2.b",          2, "3"),
    ("Q33", 2, "C-LNG", "7.L.GC.2", "7.L.GC.2.c",          2, "3"),
    ("Q34", 1, "C-LNG", "7.L.GC.2", "7.L.GC.2.d",          2, "3"),
    ("Q35", 1, "C-LNG", "7.L.GC.1", "GUM (Master Gr 6)",   1, "3"),
    ("Q36", 1, "C-LNG", "7.L.V.2",  "7.L.V.2.d",           2, "3"),
    ("Q37", 1, "C-TEC", "7.T.RA.2", "7.T.RA.2.c",          2, "2"),
    ("Q38", 1, "C-CSS", "7.T.SS.1", "7.T.SS.1.d",          2, "1"),
    ("Q39_T1", 3, "C-CSS", "Essay Trait 1", "Purpose & Organization",        3, "1"),
    ("Q39_T2", 3, "C-TEC", "Essay Trait 2", "Evidence & Elaboration",        3, "2"),
    ("Q39_T3", 2, "C-LNG", "Essay Trait 3", "Language Usage & Conventions",  3, "3"),
]

N_ITEMS = len(ITEMS)                       # 41
FIRST_ITEM_COL = 3                         # column C
LAST_ITEM_COL = FIRST_ITEM_COL + N_ITEMS - 1
CL_FIRST = get_column_letter(FIRST_ITEM_COL)          # C
CL_LAST = get_column_letter(LAST_ITEM_COL)            # AQ
COND_COL = LAST_ITEM_COL + 1                          # AR
TOTAL_COL = LAST_ITEM_COL + 2                         # AS
CL_COND = get_column_letter(COND_COL)
CL_TOTAL = get_column_letter(TOTAL_COL)

CAT_ROW, MAX_ROW, HDR_ROW = 3, 4, 5

# (code, name, max points, primary instructional focus, where to start)
CATEGORIES = [
    ("I-CSS", "Interpreting — Context, Structure & Style", 6,
     "Author's purpose and audience; perspective, tone, and credibility; how structure and figurative language shape meaning",
     "Model 'why did the author build it this way?' on short texts. Standards 7.T.C.1.a, 7.T.C.2.a-b, 7.T.SS.1.a, 7.T.SS.2.a."),
    ("I-TEC", "Interpreting — Techniques", 16,
     "Narrative, expository, argumentative, and poetic technique analysis; theme; claim and counterclaim; comparing two authors on one topic",
     "Largest category on the test (16 pts). Split by technique type before grouping — narrative gaps and argumentative gaps need different lessons."),
    ("I-PRA", "Interpreting — Periods & Movements + Research", 6,
     "Myth and adaptation analysis; comparing fiction to an account of the same material; source credibility, relevance, and bibliographic information",
     "Pair every literary text with its source myth or historical account. Teach credibility as a two-part judgment: credible AND relevant to the question."),
    ("I-LNG", "Interpreting — Language (Grammar & Vocabulary)", 6,
     "Context clues beyond the sentence; Greek and Latin roots; part of speech as a meaning clue; connotation; reading syntax",
     "Daily morphology routine plus 'read the next sentence' modeling for multiple-meaning words. Standards 7.L.V.2, 7.L.V.3, 7.L.GC.2.a."),
    ("C-CSS", "Constructing — Context, Structure & Style", 8,
     "Introductions that establish a claim; cohesive structure; transitions that name relationships; style for a target audience; conclusions that finalize",
     "Peer Revision Task routine, Trait 1 guided questions. Start with claim-writing and conclusion-writing; transitions after."),
    ("C-TEC", "Constructing — Techniques", 10,
     "Reasons that support the claim; evidence from MULTIPLE sources; elaboration in the student's own words; counterclaims; crediting sources",
     "Explicit evidence-to-elaboration instruction. Sentence stems: 'This matters because…' Reuse test items 27 and 28 as warm-ups."),
    ("C-LNG", "Constructing — Language (Grammar & Vocabulary)", 8,
     "Sentence variety; consistent verb tense; active voice; misplaced and dangling modifiers; punctuation; precise word choice",
     "Sentence combining plus targeted GUM work on rows marked M at Grades 6-7. Standards 7.L.GC.2.b-d, 7.L.GC.1."),
]

SET_CAT_FIRST = 28          # Settings row of the first category in the lookup table
LEVELS = ["Beginning Learner", "Developing Learner", "Proficient Learner", "Distinguished Learner"]

# --------------------------------------------------------------------------
# Styling
# --------------------------------------------------------------------------

FONT = "Arial"
NAVY = "1F3864"
BLUE = "2E5C8A"
LIGHT = "DCE6F1"
BAND = "F2F5FA"
YELLOW = "FFF2CC"
GREEN = "C6EFCE"
AMBER = "FFEB9C"
RED = "FFC7CE"
GREY = "E8E8E8"

H1 = Font(name=FONT, size=16, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=12, bold=True, color=NAVY)
HDR = Font(name=FONT, size=9, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
SMALL = Font(name=FONT, size=9)
SMALL_I = Font(name=FONT, size=9, italic=True, color="595959")
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HDR = PatternFill("solid", fgColor=BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW)
FILL_GREY = PatternFill("solid", fgColor=GREY)
FILL_BAND = PatternFill("solid", fgColor=BAND)

THIN = Side(style="thin", color="B0B8C4")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
CTR_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def title_bar(ws, text, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=text)
    c.font, c.fill, c.alignment = H1, FILL_TITLE, Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def subtitle(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.alignment = SMALL_I, Alignment(horizontal="left", vertical="center", indent=1)


def header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.alignment, c.border = HDR, FILL_HDR, CTR_WRAP, BOX
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def section(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.fill = H2, FILL_LIGHT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


wb = Workbook()

# ==========================================================================
# 1. START HERE
# ==========================================================================
ws = wb.active
ws.title = "START HERE"
ws.sheet_view.showGridLines = False
title_bar(ws, "Grade 7 ELA Pre/Post Assessment — Results Reporting Workbook", 8)
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 96

intro = [
    ("", ""),
    ("What this workbook does", ""),
    ("", "You enter item-level scores. It computes reporting-category profiles, achievement levels, "
         "growth statistics, item difficulty, and ready-to-teach intervention groups — automatically, "
         "for every student."),
    ("", ""),
    ("Do this, in this order", ""),
    ("1.  Roster", "Enter your students. Anything you type in a YELLOW cell is an input; everything else is a "
                   "formula and should be left alone. Row 6 contains an example student — overwrite it."),
    ("2.  Pre-Test Entry", "After administering Form A, enter the points earned on each item. "
                           "Enter 0 for a wrong answer, not a blank. Blanks are treated as 'not yet tested'."),
    ("3.  Student Report", "Read the profile. It is computed the moment you finish entering."),
    ("4.  Intervention Groups", "Sort by column F. That column is each student's primary instructional group."),
    ("5.  Post-Test Entry", "After administering Form B in the spring, enter scores the same way."),
    ("6.  Class Dashboard", "Growth, effect size, and level movement for the whole class or by period."),
    ("", ""),
    ("Cell colors", ""),
    ("Yellow", "You type here. These are the only cells you should edit."),
    ("White / grey", "Formulas. Editing one breaks the report for that student."),
    ("Blue text", "A value you may change deliberately (cut scores, thresholds) on the Settings sheet."),
    ("", ""),
    ("Score entry rules", ""),
    ("Items 1-38", "Enter the points earned: 0 or 1 for 1-point items; 0, 1, or 2 for 2-point items. "
                   "The maximum for each item is shown in row 4 of the entry sheet."),
    ("Item 39 (essay)", "Enter THREE separate trait scores: Q39_T1 (0-3), Q39_T2 (0-3), Q39_T3 (0-2). "
                        "Half-points are allowed if you double-scored and averaged."),
    ("Essay condition code", "If the essay could not be rubric-scored, leave the three trait cells BLANK and "
                             "enter a code (A-E) in the Essay Cond Code column. See the scoring guide."),
    ("", ""),
    ("Three things worth knowing before you report any of this", ""),
    ("Category scores are small", "A category rests on as few as 6 points. Treat a low category percent as a "
                                  "HYPOTHESIS to confirm against the student's actual work — never as a diagnosis "
                                  "on its own. The Student Report shows the items missed for exactly this reason."),
    ("Not every gain is real", "The workbook computes a Reliable Change Index. A gain smaller than the reliable-change "
                               "threshold on the Settings sheet may be measurement noise. Report small individual gains "
                               "as class-level movement, not as student growth."),
    ("Cut scores are provisional", "The achievement levels here were set by content review, not by a state standard-setting "
                                   "panel. They are instructional triage, NOT a prediction of a student's Georgia Milestones "
                                   "level. Replace them with locally derived cuts after your first full year."),
    ("", ""),
    ("Sheets in this workbook", ""),
    ("Settings", "Cut scores, mastery thresholds, growth target, and the intervention lookup table. All editable."),
    ("Roster", "Student names, class period, form order, and subgroup flags."),
    ("Pre-Test Entry / Post-Test Entry", "Item-level score entry for Form A and Form B."),
    ("Student Report", "One row per student: totals, levels, growth statistics, category profile, weakest areas, "
                       "and a recommended instructional focus."),
    ("Class Dashboard", "Class and period summaries, level distributions, category means, and effect size."),
    ("Item Analysis", "Item difficulty (p-values), discrimination, pre/post change, Cronbach's alpha, and flags for "
                      "items that are not working."),
    ("Intervention Groups", "Mastery heat map plus each student's primary and secondary instructional group."),
    ("Standards Coverage", "Which Grade 7 expectations this instrument does and does not measure."),
]

r = 3
for label, text in intro:
    if label and not text:
        section(ws, r, label, 3)
    else:
        if label:
            c = ws.cell(row=r, column=2, value=label)
            c.font, c.alignment = BODY_B, LEFT_WRAP
        c = ws.cell(row=r, column=3, value=text)
        c.font, c.alignment = BODY, LEFT_WRAP
        if text:
            ws.row_dimensions[r].height = max(15, 13 * (len(text) // 95 + 1))
    r += 1

ws.cell(row=r + 1, column=2, value="Source of truth").font = BODY_B
ws.cell(row=r + 1, column=3,
        value="Georgia's K-12 ELA Standards, Grade 7 (May 2023) · Georgia Milestones Grade 7 ELA Assessment Blueprint "
              "(August 2025) · Draft Grade 7 ELA Achievement Level Descriptors (October 2025) · Georgia Milestones "
              "Grade 7 Eight-Point, Three-Trait Writing Rubric (August 2026) · Grade 7 Classroom Peer Revision "
              "Guidance (February 2026).").font = SMALL_I
ws.cell(row=r + 1, column=3).alignment = LEFT_WRAP
ws.row_dimensions[r + 1].height = 40

# ==========================================================================
# 2. Settings
# ==========================================================================
ws = wb.create_sheet("Settings")
ws.sheet_view.showGridLines = False
title_bar(ws, "Settings — cut scores, thresholds, and the intervention lookup table", 6)
subtitle(ws, 2, "Blue values are editable. Every formula in this workbook reads from these cells, "
                "so changing one here updates every report.", 6)

for col, w in zip("ABCDEF", [46, 14, 12, 60, 60, 12]):
    ws.column_dimensions[col].width = w


def setting(row, label, value, note="", numfmt=None, formula=False):
    c = ws.cell(row=row, column=1, value=label)
    c.font, c.alignment = BODY, Alignment(vertical="center", indent=1)
    v = ws.cell(row=row, column=2, value=value)
    v.alignment, v.border = CTR, BOX
    if formula:
        v.font = BODY_B
        v.fill = FILL_GREY
    else:
        v.font, v.fill = INPUT_FONT, FILL_INPUT
    if numfmt:
        v.number_format = numfmt
    n = ws.cell(row=row, column=4, value=note)
    n.font, n.alignment = SMALL_I, LEFT_WRAP
    return v


section(ws, 4, "Achievement level cut scores  (raw points out of 60)", 6)
setting(5, "Developing Learner — minimum raw score", 30,
        "Below this score a student is a Beginning Learner. Default = 50% of 60.")
setting(6, "Proficient Learner — minimum raw score", 39, "Default = 65% of 60.")
setting(7, "Distinguished Learner — minimum raw score", 49, "Default = 82% of 60.")
setting(8, "Total points possible", 60, "Do not change unless you shorten the instrument.")

section(ws, 10, "Category mastery thresholds  (percent of category points)", 6)
setting(11, "Mastered — minimum percent", 0.80, "At or above: maintain through spiral review.", "0%")
setting(12, "Approaching — minimum percent", 0.60,
        "Between this and Mastered: targeted small-group instruction. Below: explicit reteaching.", "0%")

section(ws, 14, "Growth measurement", 6)
setting(15, "Individual growth target (raw points)", 9,
        "Roughly half a standard deviation. Set BELOW the reliable-change threshold so it works as a goal, "
        "not a statistical claim.")
setting(16, "Assumed reliability (Cronbach's alpha)", 0.85,
        "Replace with the value computed on the Item Analysis sheet once you have a full class of data.", "0.00")
setting(17, "Standard error of measurement (computed)",
        f"=IFERROR(ROUND(STDEV('Student Report'!$D${FIRST_ROW}:$D${LAST_ROW})*SQRT(1-$B$16),2),\"\")",
        "SD of pre-test scores x SQRT(1 - reliability). Populates once pre-test data is entered.",
        "0.00", formula=True)
setting(18, "Reliable change threshold (raw points)",
        "=IFERROR(ROUND($B$17*SQRT(2)*1.96,1),\"\")",
        "A gain at or above this is statistically reliable for an individual student (95% confidence). "
        "Smaller gains may be measurement noise.", "0.0", formula=True)

section(ws, 20, "Scoring quality (enter after double-scoring the essays)", 6)
setting(21, "Essay double-scoring exact-agreement rate", "",
        "Below 70%, do not report individual essay growth — report class-level trait movement only.", "0%")
setting(22, "Number of essays double-scored", "", "Target: at least 20% of papers.")

section(ws, 24, "Reporting category reference and intervention lookup", 6)
subtitle(ws, 25, "Column C max points replicate the Georgia Milestones Grade 7 ELA blueprint exactly. "
                 "Columns D and E feed the Recommended Focus column on the Student Report.", 6)

for j, h in enumerate(["Code", "Category", "Max pts", "Primary instructional focus", "Where to start"]):
    header_cell(ws, 27, j + 1, h)

for i, (code, name, pts, focus, start) in enumerate(CATEGORIES):
    row = SET_CAT_FIRST + i
    for j, v in enumerate([code, name, pts, focus, start]):
        c = ws.cell(row=row, column=j + 1, value=v)
        c.font, c.border = (BODY_B if j == 0 else BODY), BOX
        c.alignment = CTR if j in (0, 2) else LEFT_WRAP
        if i % 2:
            c.fill = FILL_BAND
    ws.row_dimensions[row].height = 42

tot = ws.cell(row=SET_CAT_FIRST + len(CATEGORIES), column=2, value="TOTAL")
tot.font, tot.alignment = BODY_B, Alignment(horizontal="right")
t = ws.cell(row=SET_CAT_FIRST + len(CATEGORIES), column=3,
            value=f"=SUM(C{SET_CAT_FIRST}:C{SET_CAT_FIRST+len(CATEGORIES)-1})")
t.font, t.alignment, t.border = BODY_B, CTR, BOX

lvl_row = SET_CAT_FIRST + len(CATEGORIES) + 2
section(ws, lvl_row, "Achievement level descriptors (Georgia Department of Education)", 6)
ald = [
    ("Beginning Learner", "Does not yet demonstrate proficiency in the knowledge and skills necessary at this grade "
                          "level. Needs substantial academic support to be prepared for the next grade level."),
    ("Developing Learner", "Demonstrates partial proficiency. Needs additional academic support to ensure success in "
                           "the next grade level."),
    ("Proficient Learner", "Demonstrates proficiency. Prepared for the next grade level and on track for "
                           "post-secondary readiness."),
    ("Distinguished Learner", "Demonstrates advanced proficiency. Well prepared for the next grade level and for "
                              "post-secondary readiness."),
]
for i, (lvl, desc) in enumerate(ald):
    c = ws.cell(row=lvl_row + 1 + i, column=1, value=lvl)
    c.font, c.alignment, c.border = BODY_B, Alignment(vertical="center", indent=1), BOX
    d = ws.cell(row=lvl_row + 1 + i, column=4, value=desc)
    d.font, d.alignment = SMALL, LEFT_WRAP
    ws.row_dimensions[lvl_row + 1 + i].height = 30

# ==========================================================================
# 3. Roster
# ==========================================================================
ws = wb.create_sheet("Roster")
ws.sheet_view.showGridLines = False
title_bar(ws, "Roster", 9)
subtitle(ws, 2, "Every yellow cell is an input. Row 6 is an example — overwrite it with your first student. "
                "Student order here fixes the row order on every other sheet.", 9)

roster_cols = [
    ("Student ID", 14), ("Last Name", 18), ("First Name", 16), ("Class Period", 13),
    ("Form Order", 14), ("Accommodations", 26), ("EL", 7), ("SWD", 7), ("Gifted", 8),
]
for j, (h, w) in enumerate(roster_cols):
    header_cell(ws, HDR_ROW, j + 1, h, w)

example = ["7001", "Example", "Student", "1st", "A then B", "None", "N", "N", "N"]
for j, v in enumerate(example):
    c = ws.cell(row=FIRST_ROW, column=j + 1, value=v)
    c.font, c.fill, c.border = INPUT_FONT, FILL_INPUT, BOX
    c.alignment = CTR if j >= 3 else Alignment(indent=1)

for row in range(FIRST_ROW + 1, LAST_ROW + 1):
    for j in range(9):
        c = ws.cell(row=row, column=j + 1)
        c.fill, c.border, c.font = FILL_INPUT, BOX, INPUT_FONT
        c.alignment = CTR if j >= 3 else Alignment(indent=1)

dv_order = DataValidation(type="list", formula1='"A then B,B then A"', allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws.add_data_validation(dv_order)
ws.add_data_validation(dv_yn)
dv_order.add(f"E{FIRST_ROW}:E{LAST_ROW}")
dv_yn.add(f"G{FIRST_ROW}:I{LAST_ROW}")

ws.freeze_panes = f"D{FIRST_ROW}"

n_note = ws.cell(row=3, column=1, value=f"=\"Students on roster: \"&COUNTA(C{FIRST_ROW}:C{LAST_ROW})")
n_note.font = BODY_B

# ==========================================================================
# 4 & 5. Pre-Test Entry / Post-Test Entry
# ==========================================================================


def build_entry(name, form_label):
    s = wb.create_sheet(name)
    s.sheet_view.showGridLines = False
    title_bar(s, f"{name} — {form_label}", TOTAL_COL)
    subtitle(s, 2, "Enter POINTS EARNED for each item. Enter 0 for a wrong answer — a blank means 'not tested' and "
                   "removes the student from the statistics. Row 3 shows the reporting category; row 4 shows the "
                   "maximum points for the item.", TOTAL_COL)

    s.cell(row=CAT_ROW, column=1, value="Category →").font = SMALL_I
    s.cell(row=CAT_ROW, column=1).alignment = Alignment(horizontal="right")
    s.cell(row=MAX_ROW, column=1, value="Max points →").font = SMALL_I
    s.cell(row=MAX_ROW, column=1).alignment = Alignment(horizontal="right")

    header_cell(s, HDR_ROW, 1, "Student ID", 12)
    header_cell(s, HDR_ROW, 2, "Student Name", 24)

    for i, (label, mx, cat, *_rest) in enumerate(ITEMS):
        col = FIRST_ITEM_COL + i
        c = s.cell(row=CAT_ROW, column=col, value=cat)
        c.font, c.fill, c.alignment, c.border = SMALL, FILL_LIGHT, CTR, BOX
        m = s.cell(row=MAX_ROW, column=col, value=mx)
        m.font, m.fill, m.alignment, m.border = Font(name=FONT, size=9, bold=True), FILL_GREY, CTR, BOX
        header_cell(s, HDR_ROW, col, label, 7.5 if label.startswith("Q39") else 5.5)

    header_cell(s, HDR_ROW, COND_COL, "Essay Cond Code", 11)
    header_cell(s, HDR_ROW, TOTAL_COL, "TOTAL", 9)
    s.cell(row=MAX_ROW, column=TOTAL_COL,
           value=f"=SUM({CL_FIRST}{MAX_ROW}:{CL_LAST}{MAX_ROW})").font = Font(name=FONT, size=9, bold=True)
    s.cell(row=MAX_ROW, column=TOTAL_COL).fill = FILL_GREY
    s.cell(row=MAX_ROW, column=TOTAL_COL).alignment = CTR

    dv_cond = DataValidation(type="list", formula1='"A,B,C,D,E"', allow_blank=True)
    s.add_data_validation(dv_cond)
    dv_cond.add(f"{CL_COND}{FIRST_ROW}:{CL_COND}{LAST_ROW}")

    for row in range(FIRST_ROW, LAST_ROW + 1):
        a = s.cell(row=row, column=1, value=f'=IF(Roster!$A{row}="","",Roster!$A{row})')
        a.font, a.alignment, a.border = SMALL, CTR, BOX
        b = s.cell(row=row, column=2,
                   value=f'=IF(Roster!$B{row}="","",Roster!$C{row}&" "&Roster!$B{row})')
        b.font, b.border = SMALL, BOX
        for i in range(N_ITEMS):
            col = FIRST_ITEM_COL + i
            c = s.cell(row=row, column=col)
            c.font, c.fill, c.alignment, c.border = INPUT_FONT, FILL_INPUT, CTR, BOX
        cc = s.cell(row=row, column=COND_COL)
        cc.font, cc.fill, cc.alignment, cc.border = INPUT_FONT, FILL_INPUT, CTR, BOX
        t = s.cell(row=row, column=TOTAL_COL,
                   value=f'=IF(COUNT({CL_FIRST}{row}:{CL_LAST}{row})=0,"",'
                         f'SUM({CL_FIRST}{row}:{CL_LAST}{row}))')
        t.font, t.alignment, t.border = BODY_B, CTR, BOX

    # flag any entry above the item maximum
    s.conditional_formatting.add(
        f"{CL_FIRST}{FIRST_ROW}:{CL_LAST}{LAST_ROW}",
        FormulaRule(formula=[f"AND({CL_FIRST}{FIRST_ROW}<>\"\",{CL_FIRST}{FIRST_ROW}>{CL_FIRST}${MAX_ROW})"],
                    fill=PatternFill("solid", fgColor=RED), font=Font(name=FONT, size=10, bold=True, color="9C0006")))

    s.freeze_panes = f"{CL_FIRST}{FIRST_ROW}"
    return s


build_entry("Pre-Test Entry", "FORM A")
build_entry("Post-Test Entry", "FORM B")

# ==========================================================================
# 6. Student Report
# ==========================================================================
ws = wb.create_sheet("Student Report")
ws.sheet_view.showGridLines = False

# column plan
COL_ID, COL_NAME, COL_PER = 1, 2, 3
COL_PRE_T, COL_PRE_P, COL_PRE_L = 4, 5, 6
COL_POST_T, COL_POST_P, COL_POST_L = 7, 8, 9
COL_GAIN, COL_PCTPOSS, COL_RCI, COL_REL, COL_TGT = 10, 11, 12, 13, 14
COL_PRECAT = 15                       # O .. U   (7)
COL_POSTCAT = COL_PRECAT + 7          # V .. AB
COL_CHG = COL_POSTCAT + 7             # AC .. AI
COL_WEAK1 = COL_CHG + 7               # AJ
COL_WEAK2 = COL_WEAK1 + 1             # AK
COL_NEAR = COL_WEAK2 + 1              # AL
COL_ESSAY = COL_NEAR + 1              # AM
COL_FOCUS = COL_ESSAY + 1             # AN
COL_SQ_PRE = COL_FOCUS + 2            # AP  (hidden helper)
COL_SQ_POST = COL_SQ_PRE + 1          # AQ  (hidden helper)
LAST_REP_COL = COL_FOCUS
CL_SQ_PRE = get_column_letter(COL_SQ_PRE)
CL_SQ_POST = get_column_letter(COL_SQ_POST)

title_bar(ws, "Student Report — individual profiles, growth, and recommended focus", LAST_REP_COL)
subtitle(ws, 2, "Every cell on this sheet is computed. A blank row means no scores have been entered for that "
                "student yet. 'NEAR CUT' means the student's score is within one standard error of a level "
                "boundary — for those students, trust the category profile and your own classroom evidence, "
                "not the level label.", LAST_REP_COL)

# banner row 3 groups
def banner(col_start, col_end, text, color):
    ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
    c = ws.cell(row=3, column=col_start, value=text)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = CTR


banner(COL_PRE_T, COL_PRE_L, "PRE-TEST (Form A)", "7B7B7B")
banner(COL_POST_T, COL_POST_L, "POST-TEST (Form B)", "4472A8")
banner(COL_GAIN, COL_TGT, "GROWTH", "548235")
banner(COL_PRECAT, COL_PRECAT + 6, "CATEGORY % — PRE", "7B7B7B")
banner(COL_POSTCAT, COL_POSTCAT + 6, "CATEGORY % — POST", "4472A8")
banner(COL_CHG, COL_CHG + 6, "CATEGORY CHANGE (percentage points)", "548235")
banner(COL_WEAK1, COL_FOCUS, "ACTION", "A6531C")

# category code row (row 4) — used by INDEX/MATCH for weakest-category lookup
for i, (code, *_r) in enumerate(CATEGORIES):
    for base in (COL_PRECAT, COL_POSTCAT, COL_CHG):
        c = ws.cell(row=4, column=base + i, value=code)
        c.font, c.alignment, c.fill = SMALL, CTR, FILL_LIGHT

headers = {
    COL_ID: ("Student ID", 11), COL_NAME: ("Student Name", 22), COL_PER: ("Period", 8),
    COL_PRE_T: ("Total", 8), COL_PRE_P: ("%", 8), COL_PRE_L: ("Level", 20),
    COL_POST_T: ("Total", 8), COL_POST_P: ("%", 8), COL_POST_L: ("Level", 20),
    COL_GAIN: ("Raw gain", 9), COL_PCTPOSS: ("% of possible gain", 11),
    COL_RCI: ("RCI", 8), COL_REL: ("Reliable change?", 11), COL_TGT: ("Met target?", 10),
    COL_WEAK1: ("Weakest category", 13), COL_WEAK2: ("2nd weakest", 12),
    COL_NEAR: ("Near cut?", 11), COL_ESSAY: ("Essay flag", 16), COL_FOCUS: ("Recommended focus", 62),
}
for col, (h, w) in headers.items():
    header_cell(ws, HDR_ROW, col, h, w)
for i, (code, *_r) in enumerate(CATEGORIES):
    for base in (COL_PRECAT, COL_POSTCAT, COL_CHG):
        header_cell(ws, HDR_ROW, base + i, code, 7.5)

CAT_HDR_PRE = f"$O$4:$U$4"
CAT_HDR_POST = f"$V$4:$AB$4"

for row in range(FIRST_ROW, LAST_ROW + 1):
    pre_has = f'COUNT(\'Pre-Test Entry\'!${CL_FIRST}{row}:${CL_LAST}{row})>0'
    post_has = f'COUNT(\'Post-Test Entry\'!${CL_FIRST}{row}:${CL_LAST}{row})>0'
    named = f'Roster!$B{row}<>""'

    ws.cell(row=row, column=COL_ID, value=f'=IF(Roster!$A{row}="","",Roster!$A{row})')
    ws.cell(row=row, column=COL_NAME,
            value=f'=IF({named},Roster!$C{row}&" "&Roster!$B{row},"")')
    ws.cell(row=row, column=COL_PER, value=f'=IF({named},Roster!$D{row},"")')

    ws.cell(row=row, column=COL_PRE_T,
            value=f'=IF(AND({named},{pre_has}),\'Pre-Test Entry\'!${CL_TOTAL}{row},"")')
    ws.cell(row=row, column=COL_PRE_P,
            value=f'=IFERROR(D{row}/Settings!$B$8,"")')
    ws.cell(row=row, column=COL_PRE_L,
            value=f'=IF(D{row}="","",IF(D{row}>=Settings!$B$7,"{LEVELS[3]}",'
                  f'IF(D{row}>=Settings!$B$6,"{LEVELS[2]}",'
                  f'IF(D{row}>=Settings!$B$5,"{LEVELS[1]}","{LEVELS[0]}"))))')

    ws.cell(row=row, column=COL_POST_T,
            value=f'=IF(AND({named},{post_has}),\'Post-Test Entry\'!${CL_TOTAL}{row},"")')
    ws.cell(row=row, column=COL_POST_P, value=f'=IFERROR(G{row}/Settings!$B$8,"")')
    ws.cell(row=row, column=COL_POST_L,
            value=f'=IF(G{row}="","",IF(G{row}>=Settings!$B$7,"{LEVELS[3]}",'
                  f'IF(G{row}>=Settings!$B$6,"{LEVELS[2]}",'
                  f'IF(G{row}>=Settings!$B$5,"{LEVELS[1]}","{LEVELS[0]}"))))')

    ws.cell(row=row, column=COL_GAIN, value=f'=IF(OR(D{row}="",G{row}=""),"",G{row}-D{row})')
    # "Percent of possible gain" is only meaningful for a positive gain: it asks
    # what share of the ground a student had left to cover was actually covered.
    # For a flat or declining score it is undefined, so leave it blank rather than
    # printing a nonsensical negative percentage.
    ws.cell(row=row, column=COL_PCTPOSS,
            value=f'=IFERROR(IF(OR(J{row}="",J{row}<=0),"",(G{row}-D{row})/(Settings!$B$8-D{row})),"")')
    ws.cell(row=row, column=COL_RCI,
            value=f'=IFERROR(IF(J{row}="","",(G{row}-D{row})/(Settings!$B$17*SQRT(2))),"")')
    ws.cell(row=row, column=COL_REL,
            value=f'=IF(L{row}="","",IF(ABS(L{row})>=1.96,"Yes","No"))')
    ws.cell(row=row, column=COL_TGT,
            value=f'=IF(J{row}="","",IF(J{row}>=Settings!$B$15,"Yes","No"))')

    for i in range(len(CATEGORIES)):
        set_row = SET_CAT_FIRST + i
        pc = get_column_letter(COL_PRECAT + i)
        oc = get_column_letter(COL_POSTCAT + i)
        ws.cell(row=row, column=COL_PRECAT + i,
                value=f'=IFERROR(IF($D{row}="","",'
                      f'SUMIF(\'Pre-Test Entry\'!${CL_FIRST}${CAT_ROW}:${CL_LAST}${CAT_ROW},Settings!$A${set_row},'
                      f'\'Pre-Test Entry\'!${CL_FIRST}{row}:${CL_LAST}{row})/Settings!$C${set_row}),"")')
        ws.cell(row=row, column=COL_POSTCAT + i,
                value=f'=IFERROR(IF($G{row}="","",'
                      f'SUMIF(\'Post-Test Entry\'!${CL_FIRST}${CAT_ROW}:${CL_LAST}${CAT_ROW},Settings!$A${set_row},'
                      f'\'Post-Test Entry\'!${CL_FIRST}{row}:${CL_LAST}{row})/Settings!$C${set_row}),"")')
        ws.cell(row=row, column=COL_CHG + i,
                value=f'=IF(OR({pc}{row}="",{oc}{row}=""),"",{oc}{row}-{pc}{row})')

    # weakest categories: use post if available, otherwise pre
    ws.cell(row=row, column=COL_WEAK1,
            value=f'=IFERROR(IF($G{row}<>"",INDEX({CAT_HDR_POST},MATCH(MIN($V{row}:$AB{row}),$V{row}:$AB{row},0)),'
                  f'IF($D{row}<>"",INDEX({CAT_HDR_PRE},MATCH(MIN($O{row}:$U{row}),$O{row}:$U{row},0)),"")),"")')
    ws.cell(row=row, column=COL_WEAK2,
            value=f'=IFERROR(IF($G{row}<>"",INDEX({CAT_HDR_POST},MATCH(SMALL($V{row}:$AB{row},2),$V{row}:$AB{row},0)),'
                  f'IF($D{row}<>"",INDEX({CAT_HDR_PRE},MATCH(SMALL($O{row}:$U{row},2),$O{row}:$U{row},0)),"")),"")')
    ws.cell(row=row, column=COL_NEAR,
            value=f'=IF(OR(G{row}="",Settings!$B$17=""),"",'
                  f'IF(OR(ABS(G{row}-Settings!$B$5)<=Settings!$B$17,ABS(G{row}-Settings!$B$6)<=Settings!$B$17,'
                  f'ABS(G{row}-Settings!$B$7)<=Settings!$B$17),"NEAR CUT",""))')
    ws.cell(row=row, column=COL_ESSAY,
            value=f'=IF(\'Post-Test Entry\'!${CL_COND}{row}<>"","Post essay: code "&'
                  f'\'Post-Test Entry\'!${CL_COND}{row},'
                  f'IF(\'Pre-Test Entry\'!${CL_COND}{row}<>"","Pre essay: code "&'
                  f'\'Pre-Test Entry\'!${CL_COND}{row},""))')
    ws.cell(row=row, column=COL_FOCUS,
            value=f'=IFERROR(IF($AJ{row}="","",INDEX(Settings!$D${SET_CAT_FIRST}:$D${SET_CAT_FIRST+6},'
                  f'MATCH($AJ{row},Settings!$A${SET_CAT_FIRST}:$A${SET_CAT_FIRST+6},0))),"")')

    # Hidden helpers: squared totals, so the Class Dashboard can compute a
    # conditional standard deviation without an array formula.
    ws.cell(row=row, column=COL_SQ_PRE, value=f'=IF(D{row}="","",D{row}^2)')
    ws.cell(row=row, column=COL_SQ_POST, value=f'=IF(G{row}="","",G{row}^2)')

    for col in range(1, LAST_REP_COL + 1):
        c = ws.cell(row=row, column=col)
        c.border = BOX
        c.font = SMALL
        if col in (COL_PRE_P, COL_POST_P, COL_PCTPOSS) or COL_PRECAT <= col < COL_WEAK1:
            c.number_format = "0%"
            c.alignment = CTR
        elif col == COL_RCI:
            c.number_format = "0.00"
            c.alignment = CTR
        elif col in (COL_PRE_T, COL_POST_T, COL_GAIN, COL_PER, COL_ID,
                     COL_REL, COL_TGT, COL_NEAR, COL_WEAK1, COL_WEAK2):
            c.alignment = CTR
        elif col == COL_FOCUS:
            c.alignment = LEFT_WRAP
        if col in (COL_PRE_T, COL_POST_T, COL_GAIN):
            c.font = BODY_B

rng_pre = f"O{FIRST_ROW}:U{LAST_ROW}"
rng_post = f"V{FIRST_ROW}:AB{LAST_ROW}"
for rng in (rng_pre, rng_post):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual",
                                                  formula=["Settings!$B$11"],
                                                  fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual",
                                                  formula=["Settings!$B$12"],
                                                  fill=PatternFill("solid", fgColor=AMBER)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",
                                                  formula=["Settings!$B$12"],
                                                  fill=PatternFill("solid", fgColor=RED)))
ws.conditional_formatting.add(f"AC{FIRST_ROW}:AI{LAST_ROW}",
                              ColorScaleRule(start_type="num", start_value=-0.3, start_color="F8696B",
                                             mid_type="num", mid_value=0, mid_color="FFFFFF",
                                             end_type="num", end_value=0.3, end_color="63BE7B"))
ws.conditional_formatting.add(f"AL{FIRST_ROW}:AL{LAST_ROW}",
                              CellIsRule(operator="equal", formula=['"NEAR CUT"'],
                                         fill=PatternFill("solid", fgColor=AMBER),
                                         font=Font(name=FONT, size=9, bold=True, color="7F6000")))
ws.conditional_formatting.add(f"M{FIRST_ROW}:M{LAST_ROW}",
                              CellIsRule(operator="equal", formula=['"Yes"'],
                                         fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(f"N{FIRST_ROW}:N{LAST_ROW}",
                              CellIsRule(operator="equal", formula=['"Yes"'],
                                         fill=PatternFill("solid", fgColor=GREEN)))
ws.cell(row=HDR_ROW, column=COL_SQ_PRE, value="helper: pre^2").font = SMALL_I
ws.cell(row=HDR_ROW, column=COL_SQ_POST, value="helper: post^2").font = SMALL_I
ws.column_dimensions[CL_SQ_PRE].hidden = True
ws.column_dimensions[CL_SQ_POST].hidden = True
ws.freeze_panes = f"D{FIRST_ROW}"

# ==========================================================================
# 7. Class Dashboard
# ==========================================================================
ws = wb.create_sheet("Class Dashboard")
ws.sheet_view.showGridLines = False
title_bar(ws, "Class Dashboard", 9)
subtitle(ws, 2, "Type a class period in cell B4 to filter every statistic on this sheet to that period. "
                "Leave B4 blank for all students.", 9)

for col, w in zip("ABCDEFGHI", [40, 14, 14, 14, 14, 16, 16, 16, 16]):
    ws.column_dimensions[col].width = w

c = ws.cell(row=4, column=1, value="Filter to class period (blank = all):")
c.font, c.alignment = BODY_B, Alignment(horizontal="right")
f = ws.cell(row=4, column=2, value="")
f.font, f.fill, f.border, f.alignment = INPUT_FONT, FILL_INPUT, BOX, CTR

PER = "Settings!$Z$1"     # helper holding "*" when B4 blank
wb["Settings"]["Z1"] = '=IF(\'Class Dashboard\'!$B$4="","*",\'Class Dashboard\'!$B$4)'
wb["Settings"]["Y1"] = "Helper: class-period filter used by the Class Dashboard"
wb["Settings"]["Y1"].font = SMALL_I

PRE_T = f"'Student Report'!$D${FIRST_ROW}:$D${LAST_ROW}"
POST_T = f"'Student Report'!$G${FIRST_ROW}:$G${LAST_ROW}"
PER_R = f"'Student Report'!$C${FIRST_ROW}:$C${LAST_ROW}"

section(ws, 6, "Overall performance", 9)
for j, h in enumerate(["", "Pre-Test", "Post-Test", "Change"]):
    header_cell(ws, 7, j + 1, h)

SQ_PRE = f"'Student Report'!${CL_SQ_PRE}${FIRST_ROW}:${CL_SQ_PRE}${LAST_ROW}"
SQ_POST = f"'Student Report'!${CL_SQ_POST}${FIRST_ROW}:${CL_SQ_POST}${LAST_ROW}"

# Conditional SD without an array formula:  SD = SQRT((Sum(x^2) - n*mean^2) / (n-1))
SD_PRE = (f'=IFERROR(SQRT((SUMIFS({SQ_PRE},{PER_R},{PER})-B8*B9^2)/(B8-1)),"")')
SD_POST = (f'=IFERROR(SQRT((SUMIFS({SQ_POST},{PER_R},{PER})-C8*C9^2)/(C8-1)),"")')

overall = [
    ("Students with scores",
     f'=COUNTIFS({PER_R},{PER},{PRE_T},">=0")',
     f'=COUNTIFS({PER_R},{PER},{POST_T},">=0")', None, "0"),
    ("Mean raw score (of 60)",
     f'=IFERROR(AVERAGEIFS({PRE_T},{PER_R},{PER}),"")',
     f'=IFERROR(AVERAGEIFS({POST_T},{PER_R},{PER}),"")', '=IFERROR(C9-B9,"")', "0.0"),
    ("Mean percent",
     '=IFERROR(B9/Settings!$B$8,"")', '=IFERROR(C9/Settings!$B$8,"")',
     '=IFERROR(C10-B10,"")', "0%"),
    ("Standard deviation", SD_PRE, SD_POST, None, "0.0"),
    ("Lowest score",
     f'=IFERROR(_xlfn.MINIFS({PRE_T},{PER_R},{PER}),"")',
     f'=IFERROR(_xlfn.MINIFS({POST_T},{PER_R},{PER}),"")', None, "0"),
    ("Highest score",
     f'=IFERROR(_xlfn.MAXIFS({PRE_T},{PER_R},{PER}),"")',
     f'=IFERROR(_xlfn.MAXIFS({POST_T},{PER_R},{PER}),"")', None, "0"),
]
for i, (lbl, pre, post, chg, fmt) in enumerate(overall):
    r = 8 + i
    ws.cell(row=r, column=1, value=lbl).font = BODY
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    for j, v in enumerate([pre, post, chg]):
        cell = ws.cell(row=r, column=j + 2, value=v)
        cell.font, cell.alignment, cell.border = BODY, CTR, BOX
        cell.number_format = fmt

# Cohen's d
r = 8 + len(overall) + 1
ws.cell(row=r, column=1, value="Class effect size (Cohen's d)").font = BODY_B
ws.cell(row=r, column=1).alignment = Alignment(indent=1)
d = ws.cell(row=r, column=2, value='=IFERROR((C9-B9)/SQRT((B11^2+C11^2)/2),"")')
d.font, d.alignment, d.border, d.number_format = BODY_B, CTR, BOX, "0.00"
ws.cell(row=r, column=3, value='=IFERROR(IF(B15="","",IF(B15>=0.7,"Strong year",'
                               'IF(B15>=0.4,"Solid year",IF(B15>=0.2,"Modest","Below expectation")))),"")').font = SMALL_I
ws.cell(row=r, column=4, value="Middle-grades ELA benchmarks: 0.40 solid, 0.70 strong").font = SMALL_I

section(ws, r + 2, "Achievement level distribution", 9)
lv = r + 3
for j, h in enumerate(["Achievement level", "Pre — n", "Pre — %", "Post — n", "Post — %", "Change in n"]):
    header_cell(ws, lv, j + 1, h)
for i, lvl in enumerate(LEVELS):
    rr = lv + 1 + i
    ws.cell(row=rr, column=1, value=lvl).font = BODY
    ws.cell(row=rr, column=1).alignment = Alignment(indent=1)
    ws.cell(row=rr, column=2,
            value=f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$F${FIRST_ROW}:$F${LAST_ROW},"{lvl}")')
    ws.cell(row=rr, column=3, value=f'=IFERROR(B{rr}/SUM($B${lv+1}:$B${lv+4}),"")')
    ws.cell(row=rr, column=4,
            value=f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$I${FIRST_ROW}:$I${LAST_ROW},"{lvl}")')
    ws.cell(row=rr, column=5, value=f'=IFERROR(D{rr}/SUM($D${lv+1}:$D${lv+4}),"")')
    ws.cell(row=rr, column=6, value=f'=D{rr}-B{rr}')
    for j in range(2, 7):
        cc = ws.cell(row=rr, column=j)
        cc.font, cc.alignment, cc.border = BODY, CTR, BOX
        cc.number_format = "0%" if j in (3, 5) else "0"

section(ws, lv + 6, "Reporting category performance", 9)
cat_r = lv + 7
for j, h in enumerate(["Reporting category", "Max pts", "Pre — mean %", "Post — mean %", "Change",
                       "Students 'Not Yet' (post)", "Whole-class reteach?"]):
    header_cell(ws, cat_r, j + 1, h)
for i, (code, name, pts, *_r) in enumerate(CATEGORIES):
    rr = cat_r + 1 + i
    pc = get_column_letter(COL_PRECAT + i)
    oc = get_column_letter(COL_POSTCAT + i)
    ws.cell(row=rr, column=1, value=f"{code} — {name}").font = BODY
    ws.cell(row=rr, column=1).alignment = Alignment(indent=1)
    ws.cell(row=rr, column=2, value=pts)
    ws.cell(row=rr, column=3,
            value=f'=IFERROR(AVERAGEIFS(\'Student Report\'!${pc}${FIRST_ROW}:${pc}${LAST_ROW},'
                  f'{PER_R},{PER},\'Student Report\'!${pc}${FIRST_ROW}:${pc}${LAST_ROW},">=0"),"")')
    ws.cell(row=rr, column=4,
            value=f'=IFERROR(AVERAGEIFS(\'Student Report\'!${oc}${FIRST_ROW}:${oc}${LAST_ROW},'
                  f'{PER_R},{PER},\'Student Report\'!${oc}${FIRST_ROW}:${oc}${LAST_ROW},">=0"),"")')
    ws.cell(row=rr, column=5, value=f'=IFERROR(D{rr}-C{rr},"")')
    ws.cell(row=rr, column=6,
            value=f'=COUNTIFS({PER_R},{PER},\'Student Report\'!${oc}${FIRST_ROW}:${oc}${LAST_ROW},'
                  f'"<"&Settings!$B$12,\'Student Report\'!${oc}${FIRST_ROW}:${oc}${LAST_ROW},">=0")')
    ws.cell(row=rr, column=7,
            value=f'=IF(C{rr}="","",IF(F{rr}>=0.4*MAX(1,COUNTIFS({PER_R},{PER},{POST_T},">=0")),'
                  f'"YES — reteach to whole class","No — small group"))')
    for j in range(2, 8):
        cc = ws.cell(row=rr, column=j)
        cc.font, cc.alignment, cc.border = BODY, CTR, BOX
        if j in (3, 4, 5):
            cc.number_format = "0%"
    ws.cell(row=rr, column=7).alignment = CTR_WRAP
    ws.column_dimensions["G"].width = 24

gr = cat_r + len(CATEGORIES) + 2
section(ws, gr, "Growth summary", 9)
growth_rows = [
    ("Students with both pre and post scores",
     f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$J${FIRST_ROW}:$J${LAST_ROW},">=-999")', "0"),
    ("Students meeting the growth target",
     f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$N${FIRST_ROW}:$N${LAST_ROW},"Yes")', "0"),
    ("Percent meeting the growth target",
     f'=IFERROR(B{gr+2}/B{gr+1},"")', "0%"),
    ("Students with statistically reliable change",
     f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$M${FIRST_ROW}:$M${LAST_ROW},"Yes")', "0"),
    ("Students who declined (negative gain)",
     f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$J${FIRST_ROW}:$J${LAST_ROW},"<0")', "0"),
    ("Students within one SEM of a cut score (post)",
     f'=COUNTIFS({PER_R},{PER},\'Student Report\'!$AL${FIRST_ROW}:$AL${LAST_ROW},"NEAR CUT")', "0"),
]
for i, (lbl, formula, fmt) in enumerate(growth_rows):
    rr = gr + 1 + i
    ws.cell(row=rr, column=1, value=lbl).font = BODY
    ws.cell(row=rr, column=1).alignment = Alignment(indent=1)
    cc = ws.cell(row=rr, column=2, value=formula)
    cc.font, cc.alignment, cc.border, cc.number_format = BODY_B, CTR, BOX, fmt

nr = gr + len(growth_rows) + 2
ws.cell(row=nr, column=1,
        value="Read the growth summary in this order: (1) effect size for the class, (2) percent meeting the "
              "growth target, (3) students with reliable change. A student below the reliable-change threshold "
              "has not demonstrated individually measurable growth — say so plainly rather than reporting the "
              "raw gain as if it were.").font = SMALL_I
ws.merge_cells(start_row=nr, start_column=1, end_row=nr + 1, end_column=9)
ws.cell(row=nr, column=1).alignment = LEFT_WRAP

# ==========================================================================
# 8. Item Analysis
# ==========================================================================
ws = wb.create_sheet("Item Analysis")
ws.sheet_view.showGridLines = False
title_bar(ws, "Item Analysis — difficulty, discrimination, and reliability", 13)
subtitle(ws, 2, "p-value = mean score / maximum points. Discrimination is the correlation between the item score "
                "and the student's total. An item with discrimination below 0.20 is not distinguishing stronger "
                "from weaker readers and should be reviewed before the next administration.", 13)

ia_cols = [("Item", 8), ("Max", 6), ("Category", 9), ("Standard", 14), ("Expectation", 20), ("DOK", 6),
           ("Trait", 6), ("Pre mean", 9), ("Pre p-value", 10), ("Post mean", 9), ("Post p-value", 10),
           ("Change in p", 10), ("Discrimination (pre)", 12), ("Flag", 30)]
for j, (h, w) in enumerate(ia_cols):
    header_cell(ws, HDR_ROW, j + 1, h, w)

IA_FIRST = FIRST_ROW
for i, (label, mx, cat, std, exp, dok, trait) in enumerate(ITEMS):
    r = IA_FIRST + i
    L = get_column_letter(FIRST_ITEM_COL + i)
    pre_rng = f"'Pre-Test Entry'!${L}${FIRST_ROW}:${L}${LAST_ROW}"
    post_rng = f"'Post-Test Entry'!${L}${FIRST_ROW}:${L}${LAST_ROW}"
    pre_tot = f"'Pre-Test Entry'!${CL_TOTAL}${FIRST_ROW}:${CL_TOTAL}${LAST_ROW}"

    vals = [label, mx, cat, std, exp, dok, trait,
            f'=IFERROR(AVERAGE({pre_rng}),"")',
            f'=IFERROR(H{r}/$B{r},"")',
            f'=IFERROR(AVERAGE({post_rng}),"")',
            f'=IFERROR(J{r}/$B{r},"")',
            f'=IFERROR(K{r}-I{r},"")',
            f'=IFERROR(CORREL({pre_rng},{pre_tot}),"")',
            f'=IF(I{r}="","",'
            f'IF(I{r}<0.25,"Very difficult — check key and instruction",'
            f'IF(I{r}>0.9,"Very easy — little information",'
            f'IF(AND(M{r}<>"",M{r}<0.2),"LOW DISCRIMINATION — review item","OK"))))']
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j + 1, value=v)
        c.font, c.border = SMALL, BOX
        c.alignment = LEFT_WRAP if j in (4, 13) else CTR
        if j in (8, 10, 11):
            c.number_format = "0%"
        if j in (7, 9, 12):
            c.number_format = "0.00"
        if i % 2:
            c.fill = FILL_BAND

ws.conditional_formatting.add(f"I{IA_FIRST}:I{IA_FIRST+N_ITEMS-1}",
                              ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                                             mid_type="num", mid_value=0.6, mid_color="FFEB84",
                                             end_type="num", end_value=1, end_color="63BE7B"))
ws.conditional_formatting.add(f"K{IA_FIRST}:K{IA_FIRST+N_ITEMS-1}",
                              ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                                             mid_type="num", mid_value=0.6, mid_color="FFEB84",
                                             end_type="num", end_value=1, end_color="63BE7B"))
ws.conditional_formatting.add(f"N{IA_FIRST}:N{IA_FIRST+N_ITEMS-1}",
                              CellIsRule(operator="containsText", formula=['"LOW DISCRIMINATION"'],
                                         fill=PatternFill("solid", fgColor=RED)))

rel_r = IA_FIRST + N_ITEMS + 2
section(ws, rel_r, "Reliability (Cronbach's alpha)", 13)
alpha_rows = [
    ("Number of items (k)", f"={N_ITEMS}", "0"),
    ("Sum of item variances",
     f'=IFERROR(SUM($P${IA_FIRST}:$P${IA_FIRST+N_ITEMS-1}),"")', "0.000"),
    ("Variance of total scores",
     f'=IFERROR(VARP(\'Pre-Test Entry\'!${CL_TOTAL}${FIRST_ROW}:${CL_TOTAL}${LAST_ROW}),"")', "0.000"),
    ("Cronbach's alpha (pre-test)",
     f'=IFERROR(({N_ITEMS}/({N_ITEMS}-1))*(1-B{rel_r+2}/B{rel_r+3}),"")', "0.000"),
]
for i, (lbl, formula, fmt) in enumerate(alpha_rows):
    rr = rel_r + 1 + i
    ws.cell(row=rr, column=1, value=lbl).font = BODY if i < 3 else BODY_B
    ws.cell(row=rr, column=1).alignment = Alignment(indent=1)
    c = ws.cell(row=rr, column=2, value=formula)
    c.font, c.alignment, c.border, c.number_format = (BODY if i < 3 else BODY_B), CTR, BOX, fmt

ws.cell(row=rel_r + 5, column=1,
        value="Expect alpha between 0.82 and 0.88 for a 41-item mixed-format instrument. Below 0.70, "
              "category-level reporting is not yet trustworthy for individual students — use it as a class "
              "signal only, and copy the computed alpha into Settings cell B16 so the growth statistics "
              "use your real value instead of the 0.85 default.").font = SMALL_I
ws.merge_cells(start_row=rel_r + 5, start_column=1, end_row=rel_r + 6, end_column=13)
ws.cell(row=rel_r + 5, column=1).alignment = LEFT_WRAP

# hidden helper column P: per-item variance
ws.cell(row=HDR_ROW, column=16, value="Item variance (helper)").font = SMALL_I
for i in range(N_ITEMS):
    r = IA_FIRST + i
    L = get_column_letter(FIRST_ITEM_COL + i)
    c = ws.cell(row=r, column=16,
                value=f'=IFERROR(VARP(\'Pre-Test Entry\'!${L}${FIRST_ROW}:${L}${LAST_ROW}),0)')
    c.font, c.number_format = SMALL, "0.000"
ws.column_dimensions["P"].hidden = True
ws.freeze_panes = f"A{FIRST_ROW}"

# ==========================================================================
# 9. Intervention Groups
# ==========================================================================
ws = wb.create_sheet("Intervention Groups")
ws.sheet_view.showGridLines = False
title_bar(ws, "Intervention Groups — mastery heat map and group assignments", 16)
subtitle(ws, 2, "Column F is each student's PRIMARY instructional group; column G is a secondary group. "
                "Select the table and sort by column F to form your groups. Green = Mastered (80%+), "
                "amber = Approaching (60-79%), red = Not Yet (below 60%).", 16)

section(ws, 4, "Whole-class picture — how many students are 'Not Yet' in each category?", 16)
for j, h in enumerate(["Category", "Not Yet", "Approaching", "Mastered", "Recommendation"]):
    header_cell(ws, 5, j + 1, h, [20, 11, 12, 11, 74][j])
for i, (code, name, pts, focus, start) in enumerate(CATEGORIES):
    rr = 6 + i
    oc = get_column_letter(COL_POSTCAT + i)
    pc = get_column_letter(COL_PRECAT + i)
    # use post if present, else pre
    src = f"'Student Report'!${oc}${FIRST_ROW}:${oc}${LAST_ROW}"
    src_pre = f"'Student Report'!${pc}${FIRST_ROW}:${pc}${LAST_ROW}"
    ws.cell(row=rr, column=1, value=code).font = BODY_B
    ws.cell(row=rr, column=1).alignment = CTR
    ws.cell(row=rr, column=2,
            value=f'=COUNTIFS({src},"<"&Settings!$B$12,{src},">=0")'
                  f'+COUNTIFS({src},"",{src_pre},"<"&Settings!$B$12,{src_pre},">=0")')
    ws.cell(row=rr, column=3,
            value=f'=COUNTIFS({src},">="&Settings!$B$12,{src},"<"&Settings!$B$11)'
                  f'+COUNTIFS({src},"",{src_pre},">="&Settings!$B$12,{src_pre},"<"&Settings!$B$11)')
    ws.cell(row=rr, column=4,
            value=f'=COUNTIFS({src},">="&Settings!$B$11)'
                  f'+COUNTIFS({src},"",{src_pre},">="&Settings!$B$11)')
    ws.cell(row=rr, column=5,
            value=f'=IFERROR(IF(B{rr}+C{rr}+D{rr}=0,"",IF(B{rr}/(B{rr}+C{rr}+D{rr})>=0.4,'
                  f'"WHOLE CLASS: "&Settings!$E${SET_CAT_FIRST+i},'
                  f'IF(B{rr}=0,"Maintain through spiral review",'
                  f'"SMALL GROUP ("&B{rr}&" students): "&Settings!$E${SET_CAT_FIRST+i}))),"")')
    for j in range(1, 6):
        c = ws.cell(row=rr, column=j)
        c.font, c.border = SMALL, BOX
        c.alignment = LEFT_WRAP if j == 5 else CTR
    ws.row_dimensions[rr].height = 34

ws.conditional_formatting.add(f"E6:E{5+len(CATEGORIES)}",
                              FormulaRule(formula=['LEFT($E6,11)="WHOLE CLASS"'],
                                          fill=PatternFill("solid", fgColor=RED),
                                          font=Font(name=FONT, size=9, bold=True, color="9C0006")))

IG_HDR = 6 + len(CATEGORIES) + 2
section(ws, IG_HDR - 1, "Student-level grouping — sort by column F", 16)
ig_cols = [("Student ID", 11), ("Student Name", 22), ("Period", 8), ("Post total", 10),
           ("Achievement level", 20), ("PRIMARY GROUP", 15), ("Secondary group", 14),
           ("Not Yet count", 11)]
for j, (h, w) in enumerate(ig_cols):
    header_cell(ws, IG_HDR, j + 1, h, w)
for i, (code, *_r) in enumerate(CATEGORIES):
    header_cell(ws, IG_HDR, 9 + i, code, 8)

IG_FIRST = IG_HDR + 1
IG_LAST = IG_FIRST + N_STUDENTS - 1
for k in range(N_STUDENTS):
    r = IG_FIRST + k
    sr = FIRST_ROW + k          # matching Student Report row
    ws.cell(row=r, column=1, value=f"='Student Report'!$A{sr}")
    ws.cell(row=r, column=2, value=f"='Student Report'!$B{sr}")
    ws.cell(row=r, column=3, value=f"='Student Report'!$C{sr}")
    ws.cell(row=r, column=4, value=f"='Student Report'!$G{sr}")
    ws.cell(row=r, column=5, value=f"='Student Report'!$I{sr}")
    ws.cell(row=r, column=6, value=f"='Student Report'!$AJ{sr}")
    ws.cell(row=r, column=7, value=f"='Student Report'!$AK{sr}")
    ws.cell(row=r, column=8,
            value=f'=IF($B{r}="","",COUNTIF(I{r}:O{r},"Not Yet"))')
    for i in range(len(CATEGORIES)):
        oc = get_column_letter(COL_POSTCAT + i)
        pc = get_column_letter(COL_PRECAT + i)
        ws.cell(row=r, column=9 + i,
                value=f'=IFERROR(IF(\'Student Report\'!${oc}{sr}<>"",'
                      f'IF(\'Student Report\'!${oc}{sr}>=Settings!$B$11,"Mastered",'
                      f'IF(\'Student Report\'!${oc}{sr}>=Settings!$B$12,"Approaching","Not Yet")),'
                      f'IF(\'Student Report\'!${pc}{sr}<>"",'
                      f'IF(\'Student Report\'!${pc}{sr}>=Settings!$B$11,"Mastered",'
                      f'IF(\'Student Report\'!${pc}{sr}>=Settings!$B$12,"Approaching","Not Yet")),"")),"")')
    for j in range(1, 16):
        c = ws.cell(row=r, column=j)
        c.font, c.border, c.alignment = SMALL, BOX, CTR
        if j == 2:
            c.alignment = Alignment(indent=1)
        if j == 6:
            c.font = Font(name=FONT, size=9, bold=True)

heat = f"I{IG_FIRST}:O{IG_LAST}"
ws.conditional_formatting.add(heat, CellIsRule(operator="equal", formula=['"Mastered"'],
                                               fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(heat, CellIsRule(operator="equal", formula=['"Approaching"'],
                                               fill=PatternFill("solid", fgColor=AMBER)))
ws.conditional_formatting.add(heat, CellIsRule(operator="equal", formula=['"Not Yet"'],
                                               fill=PatternFill("solid", fgColor=RED)))
ws.conditional_formatting.add(f"H{IG_FIRST}:H{IG_LAST}",
                              CellIsRule(operator="greaterThanOrEqual", formula=["4"],
                                         fill=PatternFill("solid", fgColor=RED),
                                         font=Font(name=FONT, size=9, bold=True, color="9C0006")))
ws.freeze_panes = f"C{IG_FIRST}"

# ==========================================================================
# 10. Standards Coverage
# ==========================================================================
ws = wb.create_sheet("Standards Coverage")
ws.sheet_view.showGridLines = False
title_bar(ws, "Standards Coverage — what this instrument measures, and what it does not", 5)
subtitle(ws, 2, "Read the second table before planning from this data. Roughly a quarter of the Grade 7 "
                "expectations cannot be measured in a written booklet and must be assessed through classroom "
                "performance tasks.", 5)
for col, w in zip("ABCDE", [16, 60, 10, 12, 60]):
    ws.column_dimensions[col].width = w

section(ws, 4, "Expectations assessed by this instrument", 5)
for j, h in enumerate(["Expectation", "Skill", "Items", "Category"]):
    header_cell(ws, 5, j + 1, h)

assessed = [
    ("7.L.GC.1", "Grammar, usage & mechanics (GUM: semicolons in a complex series)", "35", "C-LNG"),
    ("7.L.GC.2.a", "Apply syntax understanding to comprehend and analyze texts", "21", "I-LNG"),
    ("7.L.GC.2.b", "Sentence variety; consistent verb tense", "32, essay T3", "C-LNG"),
    ("7.L.GC.2.c", "Distinguish active and passive voice; revise for active voice", "33, essay T3", "C-LNG"),
    ("7.L.GC.2.d", "Avoid misplaced and dangling modifiers", "34, essay T3", "C-LNG"),
    ("7.L.V.2.a", "Greek and Latin roots and affixes", "14", "I-LNG"),
    ("7.L.V.2.b", "Parts of speech as a clue to meaning", "15", "I-LNG"),
    ("7.L.V.2.d", "Use parts of speech to choose precise words when writing", "36", "C-LNG"),
    ("7.L.V.3.b", "Word relationships and context clues beyond the sentence", "3", "I-LNG"),
    ("7.L.V.3.c", "Connotations of words sharing a denotation", "19", "I-LNG"),
    ("7.L.V.1.b", "Use grade-level academic vocabulary to communicate precisely", "essay T3", "C-LNG"),
    ("7.T.C.1.a", "Multiple purposes within a single text and the audiences they target", "16", "I-CSS"),
    ("7.T.C.1.c", "Construct texts for a specific purpose and audience", "essay T1", "C-CSS"),
    ("7.T.C.2.a", "Determine the prevailing perspective in a text", "20A", "I-CSS"),
    ("7.T.C.2.b", "How evidence and tone reveal perspective and affect credibility", "20B", "I-CSS"),
    ("7.T.C.2.d", "Use credible sources to research answers to questions", "essay T2", "C-TEC"),
    ("7.T.SS.1.a", "How authors modify organizational structures to achieve purposes", "11", "I-CSS"),
    ("7.T.SS.1.b", "Design texts employing structures and features", "essay T1", "C-CSS"),
    ("7.T.SS.1.c", "Varied transitions connecting ideas, sentences, and paragraphs", "30, essay T1", "C-CSS"),
    ("7.T.SS.1.d", "Multi-paragraph texts: introduction, support, conclusion", "25, 38, essay T1", "C-CSS"),
    ("7.T.SS.2.a", "How figurative and connotative language shape meaning, mood, tone", "4", "I-CSS"),
    ("7.T.SS.2.c", "Situational use of formal or informal style", "31, essay T1", "C-CSS"),
    ("7.T.T.1.a", "Narrative techniques developing plot, character, setting", "1", "I-TEC"),
    ("7.T.T.1.b", "Plot structure, conflict, and narrative devices", "2", "I-TEC"),
    ("7.T.T.1.c", "How themes are developed and expressed across texts", "5, 8", "I-TEC"),
    ("7.T.T.1.d", "Compare a fictional portrayal with an account of the same material", "10", "I-PRA"),
    ("7.T.T.2.a", "Expository techniques: main ideas, facts, statistics, text features", "12, 13", "I-TEC"),
    ("7.T.T.2.b", "How two authors on one topic emphasize different evidence", "22", "I-TEC"),
    ("7.T.T.2.d", "Apply expository techniques; elaborate on reasons", "28, essay T2", "C-TEC"),
    ("7.T.T.3.a", "Argumentative techniques: claim, evidence, counterclaim, conclusion", "17, 18", "I-TEC"),
    ("7.T.T.3.c", "Apply argumentative techniques when writing", "26, 27, 29, essay T2", "C-TEC"),
    ("7.T.T.4.a", "Poetic techniques: stanzas, rhyme, imagery, sound devices", "6, 7", "I-TEC"),
    ("7.T.PM.1.a", "Analyze a myth a modern writer has adapted; features of style and theme", "9, 10", "I-PRA"),
    ("7.T.RA.2.a", "Locate evidence; record standard bibliographic information", "24", "I-PRA"),
    ("7.T.RA.2.b", "Analyze sources for credibility and relevance", "23", "I-PRA"),
    ("7.T.RA.2.c", "Follow a standard citation format when integrating evidence", "27, 37, essay T2", "C-TEC"),
]
for i, row_vals in enumerate(assessed):
    rr = 6 + i
    for j, v in enumerate(row_vals):
        c = ws.cell(row=rr, column=j + 1, value=v)
        c.font, c.border = SMALL, BOX
        c.alignment = LEFT_WRAP if j == 1 else CTR
        if i % 2:
            c.fill = FILL_BAND

gap_r = 6 + len(assessed) + 2
section(ws, gap_r, "Expectations NOT assessed by this instrument — and how to assess them", 5)
for j, h in enumerate(["Expectation", "Skill", "", "", "How to assess it instead"]):
    header_cell(ws, gap_r + 1, j + 1, h)
gaps = [
    ("7.T.C.1.b", "Use text mode features across disciplinary texts", "",
     "Partially covered by item 13 (data table). Full coverage needs multimodal and digital texts — assess in a "
     "science or social studies reading task."),
    ("7.T.SS.2.b", "Use figurative language for intentional effect when writing", "",
     "Narrative or poetic writing task. The essay on this instrument is argumentative."),
    ("7.T.T.1.e", "Apply narrative techniques to enhance writing", "", "Narrative writing performance task."),
    ("7.T.T.4.b", "Apply poetic techniques to produce poetry", "", "Poetry writing performance task."),
    ("7.T.RA.1.a-c", "Generate research questions; conduct research; analyze from research", "",
     "Multi-day research unit with a process rubric — cannot be captured in a single sitting."),
    ("7.T.PM.1.b", "Analyze a genre of literature from a particular time period", "",
     "Extended text study with a culminating analysis."),
    ("7.L.V.1.a", "Acquire vocabulary through print, digital, and multimodal texts", "",
     "Ongoing vocabulary assessment across the year."),
    ("7.L.V.3.d / 7.L.V.3.e", "Use print and digital reference tools to clarify meaning", "",
     "Requires tool access, which conflicts with secure testing. Assess during writing conferences."),
    ("7.L.V.2.c", "Construct words from Greek and Latin roots and affixes", "",
     "Assessed only indirectly through essay word choice. Use a morphology quiz."),
    ("K-12.P.CP.1", "Collaboration", "", "Structured discussion protocol with an observation rubric."),
    ("K-12.P.CP.2", "Presentation", "", "Oral presentation with a presentation rubric."),
]
for i, row_vals in enumerate(gaps):
    rr = gap_r + 2 + i
    for j, v in enumerate(row_vals):
        c = ws.cell(row=rr, column=j + 1, value=v)
        c.font, c.border = SMALL, BOX
        c.alignment = LEFT_WRAP if j in (1, 4) else CTR
        if i % 2:
            c.fill = FILL_BAND
    ws.row_dimensions[rr].height = 30

# --------------------------------------------------------------------------
wb.save(OUT)
print(f"Wrote {OUT}")
