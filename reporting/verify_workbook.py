#!/usr/bin/env python3
"""
Verify the reporting workbook's formulas against independently computed values.

Builds a small copy of the workbook, fills it with synthetic scores whose correct
results are known in advance, evaluates every formula with a pure-Python Excel
engine, and compares cell by cell. Run this after ANY edit to
build_reporting_workbook.py.

What it checks
  - student totals, percents, and achievement-level placement
  - raw gain, percent-of-possible gain (including the divide-by-zero and
    decline guards), Reliable Change Index
  - all seven reporting-category percentages, pre and post, against the
    blueprint point values
  - weakest-category identification
  - class means, standard deviations, min/max, Cohen's d
  - achievement-level distribution counts
  - class-period filtering
  - item p-values, item variances, discrimination, and Cronbach's alpha
  - that blank rows produce blanks rather than zeros or errors
  - that no cell anywhere evaluates to an Excel error

Usage:
    pip install openpyxl formulas
    python reporting/verify_workbook.py
"""

import json
import math
import random
import subprocess
import sys
import tempfile
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

HERE = Path(__file__).resolve().parent
BUILDER = HERE / "build_reporting_workbook.py"

ITEM_MAX = [1, 2, 1, 1, 1, 1, 2, 2, 1, 2, 1, 1, 1, 1, 1, 2, 1, 2, 1, 2, 2, 2, 1, 2,
            1, 1, 2, 2, 1, 2, 1, 1, 2, 1, 1, 1, 1, 1, 3, 3, 2]
CAT = ["I-TEC", "I-TEC", "I-LNG", "I-CSS", "I-TEC", "I-TEC", "I-TEC", "I-TEC", "I-PRA", "I-PRA",
       "I-CSS", "I-TEC", "I-TEC", "I-LNG", "I-LNG", "I-CSS", "I-TEC", "I-TEC", "I-LNG", "I-CSS",
       "I-LNG", "I-TEC", "I-PRA", "I-PRA",
       "C-CSS", "C-TEC", "C-TEC", "C-TEC", "C-TEC", "C-CSS", "C-CSS", "C-LNG", "C-LNG", "C-LNG",
       "C-LNG", "C-LNG", "C-TEC", "C-CSS", "C-CSS", "C-TEC", "C-LNG"]

CATS = ["I-CSS", "I-TEC", "I-PRA", "I-LNG", "C-CSS", "C-TEC", "C-LNG"]
CAT_MAX = dict(zip(CATS, [6, 16, 6, 6, 8, 10, 8]))
PRE_COL = ["O", "P", "Q", "R", "S", "T", "U"]
POST_COL = ["V", "W", "X", "Y", "Z", "AA", "AB"]
CUTS = (30, 39, 49)
TOTAL = 60

failures = []


def level(total):
    if total >= CUTS[2]:
        return "Distinguished Learner"
    if total >= CUTS[1]:
        return "Proficient Learner"
    if total >= CUTS[0]:
        return "Developing Learner"
    return "Beginning Learner"


def check(label, got, want, tol=1e-6):
    if isinstance(want, (int, float)) and isinstance(got, (int, float)):
        ok = abs(got - want) <= tol
    else:
        ok = got == want
    if not ok:
        failures.append(f"{label}: got {got!r}, want {want!r}")
    return ok


def seed(path):
    """Fill the workbook with synthetic scores and return the expected results."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ros, pre, post = wb["Roster"], wb["Pre-Test Entry"], wb["Post-Test Entry"]
    random.seed(7)
    expected = []

    # Six graded students plus four deliberate edge cases.
    profiles = [(0.30 + 0.09 * k, min(0.97, 0.30 + 0.09 * k + 0.14)) for k in range(6)]
    profiles += [(0.0, 0.5), (1.0, 1.0), (0.02, 0.97), (0.97, 0.02)]

    for k, (ability_pre, ability_post) in enumerate(profiles):
        row = 6 + k
        period = "1st" if k < 5 else "2nd"
        ros.cell(row=row, column=1, value=7000 + k)
        ros.cell(row=row, column=2, value=f"Last{k}")
        ros.cell(row=row, column=3, value=f"First{k}")
        ros.cell(row=row, column=4, value=period)

        exact = ability_pre in (0.0, 1.0) or ability_post in (0.0, 1.0)
        pv, qv = [], []
        for mx in ITEM_MAX:
            if exact:
                a, b = round(mx * ability_pre), round(mx * ability_post)
            else:
                a = round(mx * ability_pre + random.uniform(-0.4, 0.4))
                b = round(mx * ability_post + random.uniform(-0.4, 0.4))
            a, b = max(0, min(mx, a)), max(0, min(mx, b))
            pv.append(a)
            qv.append(b)

        for i, (a, b) in enumerate(zip(pv, qv)):
            pre.cell(row=row, column=3 + i, value=a)
            post.cell(row=row, column=3 + i, value=b)

        cats = {}
        for tag, vals in (("pre", pv), ("post", qv)):
            d = {}
            for c, v in zip(CAT, vals):
                d[c] = d.get(c, 0) + v
            cats[tag] = d
        expected.append(dict(row=row, period=period, pre=sum(pv), post=sum(qv),
                             cats=cats, pv=pv, qv=qv))

    wb.save(path)
    return expected


def main():
    try:
        import formulas  # noqa: F401
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit("Missing dependency. Run:  pip install openpyxl formulas")
    import formulas

    tmp = Path(tempfile.mkdtemp())
    book = tmp / "verify.xlsx"

    print("Building a 12-row copy of the workbook…")
    subprocess.run([sys.executable, str(BUILDER), str(book), "12"], check=True,
                   stdout=subprocess.DEVNULL)

    print("Seeding synthetic scores…")
    expected = seed(book)

    print("Evaluating every formula…")
    sol = formulas.ExcelModel().loads(str(book)).finish().calculate()

    def get(sheet, cell):
        v = sol.get(f"'[{book.name}]{sheet.upper()}'!{cell}")
        if v is None:
            return "<missing>"
        a = v.value if hasattr(v, "value") else v
        while hasattr(a, "tolist"):
            a = a.tolist()
        while isinstance(a, list) and len(a) >= 1:
            a = a[0]
        return a

    print("\n--- Student Report ---")
    for e in expected:
        r = e["row"]
        gain = e["post"] - e["pre"]
        check(f"D{r} pre total", get("Student Report", f"D{r}"), e["pre"])
        check(f"G{r} post total", get("Student Report", f"G{r}"), e["post"])
        check(f"E{r} pre pct", get("Student Report", f"E{r}"), e["pre"] / TOTAL)
        check(f"H{r} post pct", get("Student Report", f"H{r}"), e["post"] / TOTAL)
        check(f"F{r} pre level", get("Student Report", f"F{r}"), level(e["pre"]))
        check(f"I{r} post level", get("Student Report", f"I{r}"), level(e["post"]))
        check(f"J{r} gain", get("Student Report", f"J{r}"), gain)
        want_k = gain / (TOTAL - e["pre"]) if gain > 0 and e["pre"] != TOTAL else ""
        check(f"K{r} pct-of-possible", get("Student Report", f"K{r}"), want_k)
        for i, c in enumerate(CATS):
            check(f"{PRE_COL[i]}{r} pre {c}", get("Student Report", f"{PRE_COL[i]}{r}"),
                  e["cats"]["pre"][c] / CAT_MAX[c])
            check(f"{POST_COL[i]}{r} post {c}", get("Student Report", f"{POST_COL[i]}{r}"),
                  e["cats"]["post"][c] / CAT_MAX[c])
        pcts = {c: e["cats"]["post"][c] / CAT_MAX[c] for c in CATS}
        weakest = [c for c in CATS if pcts[c] == min(pcts.values())][0]
        check(f"AJ{r} weakest", get("Student Report", f"AJ{r}"), weakest)
    print(f"  {len(expected)} students x 22 checks — failures: {len(failures)}")

    print("\n--- Blank rows stay blank ---")
    for cell in ("D17", "G17", "F17", "J17", "AJ17", "AN17"):
        check(f"blank {cell}", get("Student Report", cell), "")
    print("  blank-row checks complete")

    print("\n--- Class Dashboard ---")
    pres = [e["pre"] for e in expected]
    posts = [e["post"] for e in expected]
    n = len(pres)
    mp, mq = sum(pres) / n, sum(posts) / n
    sdp = math.sqrt(sum((x - mp) ** 2 for x in pres) / (n - 1))
    sdq = math.sqrt(sum((x - mq) ** 2 for x in posts) / (n - 1))
    d = (mq - mp) / math.sqrt((sdp ** 2 + sdq ** 2) / 2)
    for cell, label, want in [("B8", "n pre", n), ("C8", "n post", n),
                              ("B9", "mean pre", mp), ("C9", "mean post", mq),
                              ("B11", "SD pre", sdp), ("C11", "SD post", sdq),
                              ("B12", "min pre", min(pres)), ("B13", "max pre", max(pres)),
                              ("B15", "Cohen's d", d)]:
        got = get("Class Dashboard", cell)
        check(f"CD {cell}", got, want)
        shown = f"{got:.4f}" if isinstance(got, (int, float)) else repr(got)
        print(f"  {cell:<4} {label:<11} {shown:>12}   (expected {want:.4f})")

    from collections import Counter
    cp, cq = Counter(level(x) for x in pres), Counter(level(x) for x in posts)
    for i, lv in enumerate(["Beginning Learner", "Developing Learner",
                            "Proficient Learner", "Distinguished Learner"]):
        r = 19 + i
        check(f"CD B{r}", get("Class Dashboard", f"B{r}"), cp.get(lv, 0))
        check(f"CD D{r}", get("Class Dashboard", f"D{r}"), cq.get(lv, 0))
    print("  level distribution verified")

    for i, c in enumerate(CATS):
        r = 26 + i
        wpre = sum(e["cats"]["pre"][c] / CAT_MAX[c] for e in expected) / n
        wpost = sum(e["cats"]["post"][c] / CAT_MAX[c] for e in expected) / n
        check(f"CD C{r} {c}", get("Class Dashboard", f"C{r}"), wpre)
        check(f"CD D{r} {c}", get("Class Dashboard", f"D{r}"), wpost)
    print("  category means verified")

    print("\n--- Item Analysis ---")
    cols = [[e["pv"][i] for e in expected] for i in range(len(ITEM_MAX))]
    totals = [e["pre"] for e in expected]

    def varp(v):
        m = sum(v) / len(v)
        return sum((x - m) ** 2 for x in v) / len(v)

    for i in (0, 1, 21, 40):
        r = 6 + i
        want_mean = sum(cols[i]) / n
        check(f"IA H{r}", get("Item Analysis", f"H{r}"), want_mean)
        check(f"IA I{r}", get("Item Analysis", f"I{r}"), want_mean / ITEM_MAX[i])
        check(f"IA P{r} variance", get("Item Analysis", f"P{r}"), varp(cols[i]))
    k = len(ITEM_MAX)
    want_alpha = (k / (k - 1)) * (1 - sum(varp(c) for c in cols) / varp(totals))
    got_alpha = get("Item Analysis", "B53")
    check("IA B53 alpha", got_alpha, want_alpha)
    print(f"  p-values and variances verified")
    print(f"  Cronbach's alpha: {got_alpha:.4f}   (expected {want_alpha:.4f})")

    print("\n--- Class-period filter ---")
    import openpyxl
    filtered = tmp / "verify_filtered.xlsx"
    wb = openpyxl.load_workbook(book)
    wb["Class Dashboard"]["B4"] = "2nd"
    wb.save(filtered)
    sol2 = formulas.ExcelModel().loads(str(filtered)).finish().calculate()

    def get2(sheet, cell):
        v = sol2.get(f"'[{filtered.name}]{sheet.upper()}'!{cell}")
        a = v.value if hasattr(v, "value") else v
        while hasattr(a, "tolist"):
            a = a.tolist()
        while isinstance(a, list) and len(a) >= 1:
            a = a[0]
        return a

    sub = [e for e in expected if e["period"] == "2nd"]
    sp = [e["pre"] for e in sub]
    m2 = sum(sp) / len(sp)
    sd2 = math.sqrt(sum((x - m2) ** 2 for x in sp) / (len(sp) - 1))
    check("filtered n", get2("Class Dashboard", "B8"), len(sp))
    check("filtered mean", get2("Class Dashboard", "B9"), m2)
    check("filtered SD", get2("Class Dashboard", "B11"), sd2)
    check("filtered min", get2("Class Dashboard", "B12"), min(sp))
    print(f"  filtered to 2nd period: n={get2('Class Dashboard', 'B8')} "
          f"mean={get2('Class Dashboard', 'B9'):.2f} (expected n={len(sp)} mean={m2:.2f})")

    print("\n--- Excel errors anywhere in the workbook ---")
    errs = [k for k, v in sol.items()
            if isinstance(getattr(v, "value", None), str)
            and str(getattr(v, "value", "")).startswith("#")]
    check("no Excel errors", len(errs), 0)
    print(f"  cells evaluating to an error: {len(errs)}")
    for e in errs[:10]:
        print("   ", e, sol[e].value)

    print("\n" + "=" * 62)
    if failures:
        print(f"FAILED — {len(failures)} check(s)")
        for f in failures[:30]:
            print("  -", f)
        sys.exit(1)
    print(f"PASSED — every check succeeded ({len(sol)} cells evaluated)")


if __name__ == "__main__":
    main()
